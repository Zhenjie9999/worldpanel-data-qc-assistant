from __future__ import annotations

from html import escape
import os
from pathlib import Path
import subprocess
import tempfile
import textwrap

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .qc.issue_grouping import group_issues
from .reporting.localization import label, normalize_language


def _append_table(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, 1):
        width = max(len(header) + 4, 14)
        ws.column_dimensions[get_column_letter(index)].width = min(max(width, 18), 60)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_excel_report(
    output: Path,
    project: dict,
    run: dict,
    issues: list[dict],
    coverage: list[dict],
    ai_logs: list[dict],
    changes: list[dict] | None = None,
    matches: list[dict] | None = None,
    version_links: list[dict] | None = None,
    completion: dict | None = None,
    language: str | None = None,
) -> Path:
    language = normalize_language(language or "en")
    changes = changes or []
    matches = matches or []
    version_links = version_links or []
    wb = Workbook()
    summary = wb.active
    summary.title = label("summary_sheet", language)
    summary.append(["Worldpanel Data QC Assistant", "Local QC report"])
    summary.append(["Project", project.get("name", "")])
    summary.append(["Run ID", run.get("id", "")])
    summary.append(["Overall status", run.get("status", "Needs Review")])
    summary.append(["Issues", len(issues)])
    summary.append(["Coverage records", len(coverage)])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 72

    issue_summary = wb.create_sheet(label("issue_summary_sheet", language))
    _append_table(
        issue_summary,
        [label("severity", language), "Category", "Count", "Locations"],
        [[item["severity"], item["title"], item["count"], " | ".join(item.get("locations", []))] for item in group_issues(issues)],
    )

    qc = wb.create_sheet(label("current_qc_sheet", language))
    _append_table(
        qc,
        [
            label("severity", language),
            label("status", language),
            label("file", language),
            label("location", language),
            label("rule", language),
            label("description", language),
            label("evidence", language),
            label("recommendation", language),
            label("note", language),
        ],
        [
            [
                issue.get("severity", ""),
                issue.get("status", "pending"),
                issue.get("file_name", ""),
                issue.get("location", ""),
                issue.get("rule_id", ""),
                issue.get("description", ""),
                issue.get("evidence", ""),
                issue.get("recommendation", ""),
                issue.get("note", ""),
            ]
            for issue in issues
        ],
    )

    delta = wb.create_sheet("Changes vs Previous")
    _append_table(delta, ["Type", "File", "Location", "Before", "After"], [[c.get("type", ""), c.get("file_name", ""), c.get("location", ""), c.get("before", ""), c.get("after", "")] for c in changes])

    sources = wb.create_sheet(label("source_matching_sheet", language))
    source_rows = []
    for match in matches:
        selected = match.get("selected_candidate_index")
        if not match.get("candidates"):
            observation = match.get("observation", {})
            source_rows.append([observation.get("file_name", ""), observation.get("location", ""), observation.get("value", ""), "", "", "", "", "No candidate"])
        for index, candidate in enumerate(match.get("candidates", [])):
            observation = match.get("observation", {})
            source_rows.append(
                [
                    observation.get("file_name", ""),
                    observation.get("location", ""),
                    observation.get("value", ""),
                    candidate.get("file_name", ""),
                    candidate.get("location", ""),
                    candidate.get("value", ""),
                    candidate.get("confidence", ""),
                    "Yes" if selected == index else "",
                ]
            )
    _append_table(sources, ["Visible file", "Visible location", "Visible number", "Excel file", "Excel location", "Excel number", "Confidence", "Confirmed source"], source_rows)

    cov = wb.create_sheet(label("coverage_sheet", language))
    _append_table(
        cov,
        ["File", "Page", "Coverage %", "Numbers found", "Low confidence", "Review required", "Detail"],
        [
            [
                item.get("file_name", ""),
                item.get("page", ""),
                item.get("coverage_percent", ""),
                item.get("numbers_found", ""),
                item.get("low_confidence_count", ""),
                "Yes" if item.get("review_required") else "No",
                item.get("detail", ""),
            ]
            for item in coverage
        ],
    )

    logs = wb.create_sheet(label("ai_logs_sheet", language))
    _append_table(
        logs,
        ["Provider", "File", "Page", "Status", "Detail"],
        [[log.get("provider", ""), log.get("file_name", ""), log.get("page", ""), log.get("status", ""), log.get("detail", "")] for log in ai_logs],
    )

    links = wb.create_sheet("Version Links")
    _append_table(
        links,
        ["Current file", "Previous file", "Similarity", "Decision"],
        [[item.get("current_file_name", ""), item.get("previous_file_name", ""), item.get("similarity", ""), item.get("decision", "")] for item in version_links],
    )

    completed = wb.create_sheet("QC Completion")
    _append_table(
        completed,
        ["Completed", "User ID", "Time", "Note"],
        [["Yes", completion.get("user_id", ""), completion.get("created_at", ""), completion.get("note", "")]] if completion else [["No", "", "", ""]],
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def render_printable_summary(project: dict, run: dict, issues: list[dict], coverage: list[dict], language: str | None = None) -> str:
    language = normalize_language(language or "en")
    open_high = [i for i in issues if i.get("severity") == "High" and i.get("status", "pending") not in {"fixed", "confirmed_ok"}]
    review_pages = [c for c in coverage if c.get("review_required")]
    issue_rows = "".join(
        f"<tr><td>{escape(str(i.get('severity', '')))}</td><td>{escape(str(i.get('file_name', '')))}</td><td>{escape(str(i.get('location', '')))}</td><td>{escape(str(i.get('description', '')))}</td></tr>"
        for i in issues[:40]
    )
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>QC Summary</title>
<style>body{{font-family:Segoe UI,Arial,sans-serif;margin:36px;color:#1b2633}}table{{border-collapse:collapse;width:100%;margin-top:18px}}th,td{{border:1px solid #d9e0e7;padding:8px;text-align:left;vertical-align:top}}th{{background:#e9f0f6}}.meta{{display:flex;gap:28px}}.metric{{border-left:4px solid #1f6f8b;padding-left:10px}}</style>
</head><body>
<h1>Worldpanel Data QC Assistant</h1>
<h2>{escape(str(project.get('name', 'Project')))}</h2>
<div class="meta"><div class="metric"><b>Status</b><br>{escape(str(run.get('status', 'Needs Review')))}</div><div class="metric"><b>Issues</b><br>{len(issues)}</div><div class="metric"><b>Open High</b><br>{len(open_high)}</div><div class="metric"><b>Review pages</b><br>{len(review_pages)}</div></div>
<table><thead><tr><th>Severity</th><th>File</th><th>Location</th><th>Description</th></tr></thead><tbody>{issue_rows}</tbody></table>
<p>Use the browser print command to save this summary as PDF.</p>
</body></html>"""


def export_pdf_summary(output: Path, project: dict, run: dict, issues: list[dict], coverage: list[dict], language: str | None = None) -> Path:
    output = Path(output).resolve()
    edge = _find_edge()
    if edge:
        html = render_printable_summary(project, run, issues, coverage, language)
        output.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=output.parent) as tmp:
            html_path = Path(tmp) / "summary.html"
            profile = Path(tmp) / "edge-profile"
            html_path.write_text(html, encoding="utf-8")
            try:
                subprocess.run(
                    [
                        str(edge),
                        "--headless=new",
                        "--disable-gpu",
                        "--no-pdf-header-footer",
                        f"--user-data-dir={profile}",
                        f"--print-to-pdf={output}",
                        html_path.as_uri(),
                    ],
                    check=True,
                    capture_output=True,
                    timeout=45,
                )
                if output.exists() and output.stat().st_size:
                    return output
            except (OSError, subprocess.SubprocessError):
                pass
    return _export_basic_pdf_summary(output, project, run, issues, coverage)


def _find_edge() -> Path | None:
    candidates = [
        Path(os.environ.get("ProgramFiles(x86)", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
        Path(os.environ.get("ProgramFiles", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
    ]
    return next((path for path in candidates if path.exists()), None)


def _export_basic_pdf_summary(output: Path, project: dict, run: dict, issues: list[dict], coverage: list[dict]) -> Path:
    open_high = [i for i in issues if i.get("severity") == "High" and i.get("status", "pending") not in {"fixed", "confirmed_ok"}]
    review_pages = [c for c in coverage if c.get("review_required") and not c.get("reviewed")]
    lines = [
        "Worldpanel Data QC Assistant",
        f"Project: {project.get('name', '')}",
        f"Run ID: {run.get('id', '')}",
        f"Overall status: {run.get('status', 'Needs Review')}",
        f"Issues: {len(issues)} | Open High: {len(open_high)} | Review pages: {len(review_pages)}",
        "",
        "Issue summary",
    ]
    for issue in issues[:45]:
        text = f"[{issue.get('severity', '')}] {issue.get('file_name', '')} {issue.get('location', '')}: {issue.get('description', '')}"
        lines.extend(textwrap.wrap(text, width=105) or [""])
    if not issues:
        lines.append("No issues found by the current rules.")
    stream_lines = ["BT", "/F1 10 Tf", "48 792 Td", "14 TL"]
    for line in lines:
        safe_line = str(line).encode("latin-1", "replace").decode("latin-1").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        stream_lines.append(f"({safe_line}) Tj")
        stream_lines.append("T*")
    stream_lines.append("ET")
    stream = "\n".join(stream_lines).encode("latin-1")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    body = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, item in enumerate(objects, 1):
        offsets.append(len(body))
        body.extend(f"{index} 0 obj\n".encode("ascii"))
        body.extend(item)
        body.extend(b"\nendobj\n")
    xref = len(body)
    body.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    body.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        body.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    body.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii"))
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(body)
    return output
