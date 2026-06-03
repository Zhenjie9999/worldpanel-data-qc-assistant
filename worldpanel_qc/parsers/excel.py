from __future__ import annotations

import subprocess
import gc
import re
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook


MEASURE_PATTERN = re.compile(r"^\s*Measures?\s*=\s*(.+?)\s*$", flags=re.IGNORECASE)


def _excel_number_context(measure: str, row_label: str, column_label: str) -> str:
    return " | ".join(part for part in (measure, row_label, column_label) if part)


def _is_column_header_row(row_label: str, text_values: dict[int, str]) -> bool:
    labels_after_first_column = [value for column, value in text_values.items() if column > 1]
    return "table" in row_label.lower() or len(labels_after_first_column) >= 2


def _convert_xls(path: Path) -> tuple[Path | None, str | None]:
    libreoffice = shutil.which("soffice") or shutil.which("libreoffice")
    if libreoffice:
        target = path.with_suffix(".converted.xlsx")
        try:
            with tempfile.TemporaryDirectory(dir=path.parent) as tmp:
                output_dir = Path(tmp)
                subprocess.run(
                    [
                        libreoffice,
                        "--headless",
                        "--convert-to",
                        "xlsx",
                        "--outdir",
                        str(output_dir.resolve()),
                        str(path.resolve()),
                    ],
                    check=True,
                    capture_output=True,
                    timeout=120,
                    text=True,
                )
                converted = output_dir / f"{path.stem}.xlsx"
                if not converted.exists():
                    candidates = list(output_dir.glob("*.xlsx"))
                    if not candidates:
                        raise OSError("LibreOffice did not produce an .xlsx file.")
                    converted = candidates[0]
                shutil.copyfile(converted, target)
            return target, "Converted legacy .xls to .xlsx with LibreOffice for parsing."
        except Exception:
            pass

    target = path.with_suffix(".converted.xlsx")
    source_ps = str(path).replace("'", "''")
    target_ps = str(target).replace("'", "''")
    script = (
        "$excel = New-Object -ComObject Excel.Application; "
        "$excel.Visible = $false; "
        f"$wb = $excel.Workbooks.Open('{source_ps}'); "
        f"$wb.SaveAs('{target_ps}', 51); "
        "$wb.Close($false); $excel.Quit()"
    )
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-Command", script],
            check=True,
            capture_output=True,
            timeout=90,
            text=True,
        )
        return target, "Converted legacy .xls to .xlsx for parsing."
    except Exception:
        return None, "Unable to convert .xls automatically. Install Microsoft Excel or save the file as .xlsx."


def parse_excel(path: Path) -> dict:
    original = Path(path)
    warning = None
    parse_path = original
    if original.suffix.lower() == ".xls":
        parse_path, warning = _convert_xls(original)
        if not parse_path:
            return {
                "file_name": original.name,
                "file_type": "xls",
                "numbers": [],
                "texts": [],
                "pages": [],
                "warnings": [warning],
                "hidden_rows": [],
                "hidden_columns": [],
            }

    numbers, texts, hidden_rows, hidden_columns, formula_errors = [], [], [], [], []
    with open(parse_path, "rb") as handle:
        workbook = load_workbook(handle, data_only=False, read_only=False)
        for ws in workbook.worksheets:
            measure = ""
            column_labels = {}
            for index, dimension in ws.row_dimensions.items():
                if dimension.hidden:
                    hidden_rows.append(f"{ws.title}!{index}")
            for label, dimension in ws.column_dimensions.items():
                if dimension.hidden:
                    hidden_columns.append(f"{ws.title}!{label}")
            for row in ws.iter_rows():
                row_label = str(ws.cell(row[0].row, 1).value or "").strip()
                measure_match = MEASURE_PATTERN.match(row_label)
                if measure_match:
                    measure = measure_match.group(1).strip()
                text_values = {
                    cell.column: str(cell.value).strip()
                    for cell in row
                    if isinstance(cell.value, str) and str(cell.value).strip()
                }
                if _is_column_header_row(row_label, text_values):
                    column_labels = text_values
                for cell in row:
                    value = cell.value
                    location = f"{ws.title}!{cell.coordinate}"
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        column_label = column_labels.get(cell.column, "")
                        numbers.append(
                            {
                                "value": float(value),
                                "context": _excel_number_context(measure, row_label, column_label),
                                "location": location,
                                "file_name": original.name,
                                "is_percent": "%" in str(cell.number_format),
                                "measure": measure,
                                "row_label": row_label,
                                "column_label": column_label,
                            }
                        )
                    elif isinstance(value, str) and value.strip():
                        texts.append({"text": value.strip(), "location": location})
                    if isinstance(value, str) and value.startswith("#"):
                        formula_errors.append(location)
        workbook.close()
        del workbook
    gc.collect()
    warnings = [warning] if warning else []
    if parse_path != original and parse_path.exists():
        try:
            parse_path.unlink()
        except OSError:
            pass
    return {
        "file_name": original.name,
        "file_type": original.suffix.lower().lstrip("."),
        "numbers": numbers,
        "texts": texts,
        "pages": [],
        "warnings": warnings,
        "hidden_rows": hidden_rows,
        "hidden_columns": hidden_columns,
        "formula_errors": formula_errors,
    }
