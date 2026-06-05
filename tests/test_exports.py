import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from worldpanel_qc.exports import export_excel_report, export_pdf_summary


class ExportTests(unittest.TestCase):
    def test_excel_export_contains_required_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "report.xlsx"
            export_excel_report(
                output,
                {"name": "Zespri PanelVoice"},
                {"id": 1, "status": "Needs Review"},
                [{"severity": "High", "description": "Placeholder", "file_name": "deck.pptx", "location": "Slide 20", "status": "pending"}],
                [{"file_name": "deck.pptx", "page": 20, "coverage_percent": 70, "numbers_found": 1, "low_confidence_count": 1, "review_required": True}],
                [],
            )
            wb = load_workbook(output, read_only=True)
            self.assertEqual(
                wb.sheetnames,
                [
                    "Summary",
                    "Issue Summary",
                    "Current File QC",
                    "Changes vs Previous",
                    "Source Matching Concerns",
                    "Coverage",
                    "AI Logs",
                    "Version Links",
                    "QC Completion",
                ],
            )
            wb.close()

    def test_excel_export_uses_selected_language_for_sheet_names(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "report-zh.xlsx"
            export_excel_report(
                output,
                {"name": "Zespri PanelVoice"},
                {"id": 1, "status": "Needs Review"},
                [],
                [],
                [],
                language="zh",
            )
            wb = load_workbook(output, read_only=True)
            self.assertIn("摘要", wb.sheetnames)
            self.assertIn("当前文件QC", wb.sheetnames)
            wb.close()

    def test_pdf_summary_is_a_downloadable_pdf(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "summary.pdf"
            export_pdf_summary(
                output,
                {"name": "Zespri PanelVoice"},
                {"id": 1, "status": "Ready for Delivery"},
                [],
                [],
            )
            self.assertTrue(output.read_bytes().startswith(b"%PDF-"))


if __name__ == "__main__":
    unittest.main()
